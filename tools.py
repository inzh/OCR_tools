from configparser import ConfigParser
import openpyxl
def configSeting(ak, sk):
    conf = ConfigParser()
    conf.add_section("config")
    conf.set('config', 'Api Key', ak)
    conf.set('config', 'Secret Key', sk)
    with open('config.ini', 'w') as fw:
        conf.write(fw)

def getAK():
    conf = ConfigParser()
    file = 'config.ini'
    conf.read(file)
    ak = conf.get("config","api key")
    return ak

def getSK():
    conf = ConfigParser()
    file = 'config.ini'
    conf.read(file)
    sk = conf.get("config", "secret key")
    return sk

def create_sheet():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "OCR数据"
    sheet.cell(1, 1, value="姓名")
    sheet.cell(1, 2, value="性别")
    sheet.cell(1, 3, value="民族")
    sheet.cell(1, 4, value="出生年月")
    sheet.cell(1, 5, value="身份证号码")
    sheet.cell(1, 6, value="地址")
    sheet.cell(1, 7, value="文件名")
    workbook.save("ocr.xlsx")
    # print("xlsx格式表格写入数据成功！")

def set_sheet():
    wb = openpyxl.load_workbook('ocr.xlsx')
    # print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    # 调整列宽
    ws.column_dimensions['D'].width = 20.0
    ws.column_dimensions['E'].width = 40.0
    ws.column_dimensions['F'].width = 80.0
    ws.column_dimensions['G'].width = 100.0
    wb.save('ocr.xlsx')

def sheet_append(name, sex, nation, birth, num, address,filename):
    wb = openpyxl.load_workbook('ocr.xlsx')
    ws = wb[wb.sheetnames[0]]
    data = [name, sex, nation, birth, num, address, filename]
    ws.append(data)
    wb.save('ocr.xlsx')